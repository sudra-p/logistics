"""
Tests for report export endpoints (CSV and Excel).
"""

import csv
import datetime
import io

import pytest
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APIClient

from bookings.models import Booking, Container
from master_data.models import (
    Client,
    Commodity,
    ContainerType,
    Port,
    ShippingLine,
    Vessel,
)
from reports.exporters import MAX_EXPORT_ROWS

User = get_user_model()

PENDING_DO_EXPORT_URL = '/api/reports/pending-do/export/'
MASTER_EXPORT_URL = '/api/reports/master/export/'


@pytest.fixture
def ops_user(db):
    """Create an Operations group user."""
    user = User.objects.create_user(username='ops_export', password='testpass123')
    group, _ = Group.objects.get_or_create(name='Operations')
    user.groups.add(group)
    return user


@pytest.fixture
def sales_user(db):
    """Create a Sales group user (no modify permission)."""
    user = User.objects.create_user(username='sales_export', password='testpass123')
    group, _ = Group.objects.get_or_create(name='Sales')
    user.groups.add(group)
    return user


@pytest.fixture
def master_data(db):
    """Create required master data for bookings."""
    shipping_line = ShippingLine.objects.create(name='Maersk', code='MAEU')
    pol = Port.objects.create(name='Mumbai Port', code='INMUN', country='India')
    pod = Port.objects.create(name='Rotterdam Port', code='NLRTM', country='Netherlands')
    client = Client.objects.create(name='Test Client', email='client@test.com')
    commodity = Commodity.objects.create(name='Electronics', hs_code='8542')
    container_type = ContainerType.objects.create(name='Standard', code='GP')
    vessel = Vessel.objects.create(name='Ever Given', shipping_line=shipping_line)
    return {
        'shipping_line': shipping_line,
        'pol': pol,
        'pod': pod,
        'client': client,
        'commodity': commodity,
        'container_type': container_type,
        'vessel': vessel,
    }


@pytest.fixture
def api_client():
    return APIClient()


def _create_booking(master_data, ops_user, **overrides):
    """Helper to create a booking with defaults."""
    today = timezone.now().date()
    defaults = {
        'booking_date': today - datetime.timedelta(days=5),
        'booking_validity_date': today + datetime.timedelta(days=10),
        'forwarding_window_start': today + datetime.timedelta(days=1),
        'forwarding_window_end': today + datetime.timedelta(days=7),
        'shipping_line': master_data['shipping_line'],
        'pol': master_data['pol'],
        'pod': master_data['pod'],
        'client': master_data['client'],
        'commodity': master_data['commodity'],
        'cargo_type': 'FCL',
        'shipment_type': 'Direct',
        'stuffing_type': 'Factory',
        'status': Booking.Status.PENDING,
        'created_by': ops_user,
    }
    defaults.update(overrides)
    return Booking.objects.create(**defaults)


class TestPendingDOExportCSV:
    """Tests for Pending DO CSV export."""

    def test_csv_export_has_correct_headers(self, api_client, ops_user, master_data):
        """CSV export contains all expected column headers."""
        api_client.force_authenticate(user=ops_user)
        _create_booking(master_data, ops_user)

        response = api_client.get(PENDING_DO_EXPORT_URL, {'format': 'csv'})

        assert response.status_code == status.HTTP_200_OK
        assert response['Content-Type'] == 'text/csv'
        assert 'attachment; filename="pending_do_report.csv"' in response['Content-Disposition']

        content = response.content.decode('utf-8')
        reader = csv.reader(io.StringIO(content))
        headers = next(reader)

        expected_headers = [
            'Booking Reference',
            'Client Name',
            'Vessel/Voyage',
            'POL',
            'POD',
            'ETD',
            'ETA',
            'Status',
            'Shipping Line',
            'Container Count',
            'Booking Date',
        ]
        assert headers == expected_headers

    def test_csv_export_includes_data(self, api_client, ops_user, master_data):
        """CSV export includes booking data rows."""
        api_client.force_authenticate(user=ops_user)
        booking = _create_booking(
            master_data, ops_user, vessel=master_data['vessel'], voyage='V100'
        )
        Container.objects.create(
            booking=booking,
            container_type=master_data['container_type'],
            container_size='20FT',
            container_count=3,
        )

        response = api_client.get(PENDING_DO_EXPORT_URL, {'format': 'csv'})

        assert response.status_code == status.HTTP_200_OK
        content = response.content.decode('utf-8')
        reader = csv.reader(io.StringIO(content))
        next(reader)  # skip headers
        row = next(reader)

        assert row[0] == booking.job_number
        assert row[1] == 'Test Client'
        assert 'Ever Given' in row[2]
        assert 'V100' in row[2]
        assert row[3] == 'Mumbai Port'
        assert row[4] == 'Rotterdam Port'
        assert row[7] == 'Pending'
        assert row[8] == 'Maersk'
        assert row[9] == '3'

    def test_csv_export_filters_applied(self, api_client, ops_user, master_data):
        """CSV export respects query parameter filters."""
        api_client.force_authenticate(user=ops_user)
        client_2 = Client.objects.create(name='Another Client', email='c2@test.com')

        b1 = _create_booking(master_data, ops_user, client=master_data['client'])
        b2 = _create_booking(master_data, ops_user, client=client_2)

        response = api_client.get(
            PENDING_DO_EXPORT_URL,
            {'format': 'csv', 'client': master_data['client'].pk},
        )

        assert response.status_code == status.HTTP_200_OK
        content = response.content.decode('utf-8')
        reader = csv.reader(io.StringIO(content))
        next(reader)  # skip headers
        rows = list(reader)

        job_numbers = [row[0] for row in rows]
        assert b1.job_number in job_numbers
        assert b2.job_number not in job_numbers


class TestPendingDOExportExcel:
    """Tests for Pending DO Excel export."""

    def test_excel_export_produces_valid_xlsx(self, api_client, ops_user, master_data):
        """Excel export produces a valid xlsx file with correct headers."""
        api_client.force_authenticate(user=ops_user)
        _create_booking(master_data, ops_user)

        response = api_client.get(PENDING_DO_EXPORT_URL, {'format': 'excel'})

        assert response.status_code == status.HTTP_200_OK
        assert (
            response['Content-Type']
            == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        assert 'attachment; filename="pending_do_report.xlsx"' in response['Content-Disposition']

        # Parse the xlsx content
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        expected_headers = [
            'Booking Reference',
            'Client Name',
            'Vessel/Voyage',
            'POL',
            'POD',
            'ETD',
            'ETA',
            'Status',
            'Shipping Line',
            'Container Count',
            'Booking Date',
        ]
        assert headers == expected_headers

    def test_excel_export_includes_data(self, api_client, ops_user, master_data):
        """Excel export includes booking data rows."""
        api_client.force_authenticate(user=ops_user)
        booking = _create_booking(master_data, ops_user)

        response = api_client.get(PENDING_DO_EXPORT_URL, {'format': 'excel'})

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Row 2 is first data row (row 1 is header)
        assert ws.cell(row=2, column=1).value == booking.job_number
        assert ws.cell(row=2, column=2).value == 'Test Client'

    def test_excel_export_bold_headers(self, api_client, ops_user, master_data):
        """Excel export has bold header row."""
        api_client.force_authenticate(user=ops_user)
        _create_booking(master_data, ops_user)

        response = api_client.get(PENDING_DO_EXPORT_URL, {'format': 'excel'})

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        for cell in ws[1]:
            assert cell.font.bold is True


class TestMasterExportCSV:
    """Tests for Master report CSV export."""

    def test_csv_export_has_correct_headers(self, api_client, ops_user, master_data):
        """Master CSV export has Created Date instead of Booking Date."""
        api_client.force_authenticate(user=ops_user)
        _create_booking(master_data, ops_user)

        response = api_client.get(MASTER_EXPORT_URL, {'format': 'csv'})

        assert response.status_code == status.HTTP_200_OK
        content = response.content.decode('utf-8')
        reader = csv.reader(io.StringIO(content))
        headers = next(reader)

        expected_headers = [
            'Booking Reference',
            'Client Name',
            'Vessel/Voyage',
            'POL',
            'POD',
            'ETD',
            'ETA',
            'Status',
            'Shipping Line',
            'Container Count',
            'Created Date',
        ]
        assert headers == expected_headers


class TestMasterExportExcel:
    """Tests for Master report Excel export."""

    def test_excel_export_produces_valid_xlsx(self, api_client, ops_user, master_data):
        """Master Excel export produces a valid xlsx file."""
        api_client.force_authenticate(user=ops_user)
        _create_booking(master_data, ops_user)

        response = api_client.get(MASTER_EXPORT_URL, {'format': 'excel'})

        assert response.status_code == status.HTTP_200_OK
        assert (
            response['Content-Type']
            == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        assert ws.cell(row=1, column=11).value == 'Created Date'


class TestExportFormatValidation:
    """Tests for format query parameter validation."""

    def test_missing_format_returns_400(self, api_client, ops_user, master_data):
        """Missing format query parameter returns 400."""
        api_client.force_authenticate(user=ops_user)

        response = api_client.get(PENDING_DO_EXPORT_URL)

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert 'format' in response.data['error'].lower()

    def test_invalid_format_returns_400(self, api_client, ops_user, master_data):
        """Invalid format value returns 400."""
        api_client.force_authenticate(user=ops_user)

        response = api_client.get(PENDING_DO_EXPORT_URL, {'format': 'pdf'})

        assert response.status_code == status.HTTP_400_BAD_REQUEST

    def test_master_missing_format_returns_400(self, api_client, ops_user, master_data):
        """Master export also requires format param."""
        api_client.force_authenticate(user=ops_user)

        response = api_client.get(MASTER_EXPORT_URL)

        assert response.status_code == status.HTTP_400_BAD_REQUEST


class TestExportPermissions:
    """Tests for permission enforcement on export endpoints."""

    def test_unauthenticated_denied_pending_do(self, api_client):
        """Unauthenticated requests to Pending DO export are denied."""
        response = api_client.get(PENDING_DO_EXPORT_URL, {'format': 'csv'})
        assert response.status_code == status.HTTP_401_UNAUTHORIZED

    def test_unauthenticated_denied_master(self, api_client):
        """Unauthenticated requests to Master export are denied."""
        response = api_client.get(MASTER_EXPORT_URL, {'format': 'csv'})
        assert response.status_code == status.HTTP_401_UNAUTHORIZED

    def test_sales_user_denied_pending_do(self, api_client, sales_user):
        """Sales users cannot access Pending DO export."""
        api_client.force_authenticate(user=sales_user)
        response = api_client.get(PENDING_DO_EXPORT_URL, {'format': 'csv'})
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_sales_user_denied_master(self, api_client, sales_user):
        """Sales users cannot access Master export."""
        api_client.force_authenticate(user=sales_user)
        response = api_client.get(MASTER_EXPORT_URL, {'format': 'csv'})
        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_ops_user_allowed_pending_do(self, api_client, ops_user):
        """Operations users can access Pending DO export."""
        api_client.force_authenticate(user=ops_user)
        response = api_client.get(PENDING_DO_EXPORT_URL, {'format': 'csv'})
        assert response.status_code == status.HTTP_200_OK

    def test_ops_user_allowed_master(self, api_client, ops_user):
        """Operations users can access Master export."""
        api_client.force_authenticate(user=ops_user)
        response = api_client.get(MASTER_EXPORT_URL, {'format': 'csv'})
        assert response.status_code == status.HTTP_200_OK


class TestExportRowCap:
    """Tests for the 50,000 row export cap."""

    def test_max_export_rows_constant(self):
        """MAX_EXPORT_ROWS is 50,000."""
        assert MAX_EXPORT_ROWS == 50_000
